from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def load_txt(path: str | Path) -> list[dict[str, str]]:
    lines = Path(path).read_text(encoding="utf-8-sig").splitlines()
    return [{"nombre": line.strip()} for line in lines if line.strip()]


def rows_from_pasted_text(text: str) -> list[dict[str, str]]:
    rows = [line.strip() for line in text.splitlines() if line.strip()]
    if not rows:
        return []
    if any("\t" in row for row in rows):
        split_rows = [row.split("\t") for row in rows]
        width = max(len(row) for row in split_rows)
        headers = ["nombre"] + [f"columna_{i}" for i in range(2, width + 1)]
        return [_row_to_dict(headers, row) for row in split_rows]
    return [{"nombre": row} for row in rows]


def load_csv(path: str | Path) -> list[dict[str, str]]:
    sample = Path(path).read_text(encoding="utf-8-sig", errors="replace")[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    has_header = csv.Sniffer().has_header(sample)
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, dialect)
        rows = [[cell.strip() for cell in row] for row in reader if any(cell.strip() for cell in row)]
    return normalize_rows(rows, has_header=has_header)


def load_xlsx(path: str | Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if cell is None else str(cell).strip() for cell in row]
        if any(values):
            rows.append(values)
    wb.close()
    return normalize_rows(rows, has_header=_looks_like_header(rows))


def load_data_file(path: str | Path) -> list[dict[str, str]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".txt":
        return load_txt(path)
    if suffix == ".csv":
        return load_csv(path)
    if suffix == ".xlsx":
        return load_xlsx(path)
    raise ValueError(f"Formato no soportado: {suffix}")


def normalize_rows(rows: list[list[str]], has_header: bool | None = None) -> list[dict[str, str]]:
    if not rows:
        return []
    width = max(len(row) for row in rows)
    if has_header is None:
        has_header = _looks_like_header(rows)
    if has_header:
        headers = [_normalize_header(cell, i + 1) for i, cell in enumerate(_pad(rows[0], width))]
        data_rows = rows[1:]
    else:
        headers = ["nombre"] if width == 1 else [f"columna_{i}" for i in range(1, width + 1)]
        data_rows = rows
    return [_row_to_dict(headers, row) for row in data_rows if any(str(cell).strip() for cell in row)]


def _row_to_dict(headers: Iterable[str], row: list[str]) -> dict[str, str]:
    padded = _pad(row, len(list(headers)) if not isinstance(headers, list) else len(headers))
    return {header: str(padded[index]).strip() for index, header in enumerate(headers)}


def _pad(row: list[str], width: int) -> list[str]:
    return row + [""] * max(0, width - len(row))


def _normalize_header(value: str, index: int) -> str:
    header = str(value).strip().lower().replace(" ", "_")
    return header or f"columna_{index}"


def _looks_like_header(rows: list[list[str]]) -> bool:
    if len(rows) < 2:
        return False
    first = [cell.strip().lower() for cell in rows[0]]
    if not first or not any(first):
        return False
    common = {"nombre", "apellido", "curso", "fecha", "alumno", "numero", "nro", "grado"}
    return any(cell in common for cell in first)
